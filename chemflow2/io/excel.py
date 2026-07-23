"""ストリーム結果の Excel(.xlsx) 出力。

元 chemflow は win32com で「起動中の Excel ブック」に書き込む Windows 専用方式だったが、
chemflow2 ではスタンドアロンの .xlsx ファイルを生成する（クロスプラットフォーム）。

basis を指定して mol/h・g/h・NL/h と各分率を basis ごとのブロックで出せる（既定 "mol"）。
openpyxl はオプション依存: `pip install chemflow2[excel]`
"""

from __future__ import annotations

import numpy as np

from chemflow2.core.stream import Stream
from chemflow2.io.table import (
    _UNITS,
    _all_formulas,
    _as_bases,
    _formula_maps,
    _ordered,
    _stream_values,
    _total_mass,
)


def to_excel(streams: list[Stream], path: str, *, sheet: str = "Streams", basis="mol") -> None:
    """成分 × ストリームの流量表を .xlsx として書き出す。

    レイアウト:
        ヘッダ（ストリーム名 / T / P / Phase）
        basis ごとのブロック（単位行・成分行・合計行）
        質量閉包の確認行（mass ブロックが無いとき）

    basis: "mol"（既定）/"mass"/"normal_volume"/"mole_frac"/"mass_frac"/"volume_frac"、
    もしくはそれらのリスト。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:  # pragma: no cover
        raise ImportError("openpyxl が必要です: pip install chemflow2[excel]") from e

    bases = _as_bases(basis)
    streams = _ordered(streams)
    formulas = _all_formulas(streams)
    mw, nv = _formula_maps(formulas)
    names = [s.name or f"S{i}" for i, s in enumerate(streams)]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    bold = Font(bold=True)

    def _cond(s: Stream, attr: str):
        v = getattr(s.condition, attr, None)
        return "" if v is None else v

    # --- ヘッダ ---
    ws.append(["Component", "MW"] + names)
    for c in ws[ws.max_row]:
        c.font = bold
    ws.append(["T [°C]", ""] + [_cond(s, "T") for s in streams])
    ws.append(["P", ""] + [_cond(s, "P") for s in streams])
    ws.append(["Phase", ""] + [_cond(s, "phase") for s in streams])

    # --- basis ごとのブロック ---
    for b in bases:
        unit = _UNITS[b]
        ws.append([])
        ws.append([f"[{unit}]", ""] + ["" for _ in streams])
        cols = [_stream_values(s, formulas, mw, nv, b) for s in streams]
        for i, f in enumerate(formulas):
            ws.append([f, round(mw[i], 4)] + [float(col[i]) for col in cols])
        total_r = ws.max_row + 1
        ws.append([f"total [{unit}]", ""] + [float(col.sum()) for col in cols])
        ws.cell(row=total_r, column=1).font = bold

    # --- 質量閉包（mass ブロックが無いとき）---
    if "mass" not in bases:
        ws.append([])
        r = ws.max_row + 1
        ws.append(["total mass [g/h]", ""] + [_total_mass(s) for s in streams])
        ws.cell(row=r, column=1).font = bold

    # 列幅を軽く整える
    ws.column_dimensions["A"].width = 18
    for i in range(len(names)):
        ws.column_dimensions[chr(ord("C") + i)].width = 12

    wb.save(path)
