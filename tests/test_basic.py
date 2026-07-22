"""基本動作の最小テスト。

    PYTHONPATH=. python3 -m pytest tests/ -q
"""

import numpy as np
import pytest

from chemflow2 import (
    Expr,
    Mixer,
    Problem,
    Reaction,
    ReactionError,
    Reactor,
    Separator,
    Splitter,
    Stream,
    StreamCondition,
    SolveError,
    generate_mermaid,
    parse_pressure,
)


def test_component_mw():
    s = Stream(["H2O"])
    assert abs(s.components[0].mw - 18.015) < 0.01


def test_mixer_balance():
    a = Stream(["N2", "H2"], flows={"N2": 1, "H2": 3})
    b = Stream(["N2", "H2"], flows={"N2": 2, "H2": 1})
    out = Stream(["N2", "H2"], name="out")
    p = Problem([a, b, out], [Mixer([a, b], out)])
    sol = p.solve()
    assert sol.success
    assert out.flow_of("N2") == pytest.approx(3.0)
    assert out.flow_of("H2") == pytest.approx(4.0)


def test_reactor_conversion():
    feed = Stream(["N2O4", "NO2"], flows={"N2O4": 10, "NO2": 0})
    out = Stream(["N2O4", "NO2"], name="out")
    rxn = Reaction({"N2O4": -1, "NO2": 2})
    p = Problem([feed, out], [Reactor(feed, out, [rxn], key_component="N2O4", conversion=0.5)])
    sol = p.solve()
    assert sol.success
    assert out.flow_of("N2O4") == pytest.approx(5.0)
    assert out.flow_of("NO2") == pytest.approx(10.0)


def test_recycle_converges_and_balances():
    comps = ["N2O4", "NO2"]
    feed = Stream(comps, flows={"N2O4": 10, "NO2": 0})
    rec = Stream(comps, name="rec")
    mixed = Stream(comps, name="mixed")
    rout = Stream(comps, name="rout")
    prod = Stream(comps, name="prod")
    rxn = Reaction({"N2O4": -1, "NO2": 2})
    p = Problem(
        [feed, rec, mixed, rout, prod],
        [
            Mixer([feed, rec], mixed),
            Reactor(mixed, rout, [rxn], key_component="N2O4", conversion=0.5),
            Separator(rout, [prod, rec]),
        ],
    )
    p.constrain(Expr(lambda: rec.molar_flows - 0.3 * rout.molar_flows))
    sol = p.solve(bounds=(0, np.inf))
    assert sol.success
    # 元素収支: 入った N は出た N（Product）と等しい（N2O4=2N, NO2=1N）
    n_in = 2 * feed.flow_of("N2O4") + feed.flow_of("NO2")
    n_out = 2 * prod.flow_of("N2O4") + prod.flow_of("NO2")
    assert n_in == pytest.approx(n_out)


def test_reaction_atom_balance_ok():
    # 2 CO + 4 H2 -> CH3OCH3 + H2O は原子保存する
    Reaction({"CO": -2, "H2": -4, "CH3OCH3": 1, "H2O": 1})


def test_reaction_atom_balance_bad_raises():
    # H2O の係数をタイポ（原子が消える）
    with pytest.raises(ReactionError):
        Reaction({"CO": -2, "H2": -4, "CH3OCH3": 1})


def test_selectivity_sum_must_be_one():
    feed = Stream(["N2O4", "NO2"], flows={"N2O4": 10, "NO2": 0})
    out = Stream(["N2O4", "NO2"], name="out")
    rxn = Reaction({"N2O4": -1, "NO2": 2})
    with pytest.raises(ValueError):
        Reactor(feed, out, [rxn, rxn], key_component="N2O4",
                conversion=0.5, selectivities=[0.7, 0.7])


def test_flow_expr_constraint():
    feed = Stream(["N2O4", "NO2"], flows={"N2O4": 10, "NO2": 5})
    a = Stream(["N2O4", "NO2"], name="a")
    b = Stream(["N2O4", "NO2"], name="b")
    p = Problem([feed, a, b], [Separator(feed, [a, b])])
    p.constrain(b.flow_expr("N2O4"), 0.0)
    p.constrain(b.flow_expr("NO2"), 2.0)
    sol = p.solve()
    assert sol.success
    assert a.flow_of("N2O4") == pytest.approx(10.0)
    assert a.flow_of("NO2") == pytest.approx(3.0)


def test_splitter_ratio_and_composition():
    feed = Stream(["N2O4", "NO2"], flows={"N2O4": 6, "NO2": 4})
    a = Stream(["N2O4", "NO2"], name="a")
    b = Stream(["N2O4", "NO2"], name="b")
    p = Problem([feed, a, b], [Splitter(feed, [a, b], ratios=[0.7, 0.3])])
    sol = p.solve()
    assert sol.success
    # 比率どおり & 組成は入口と同一
    assert a.flow_of("N2O4") == pytest.approx(4.2)
    assert a.flow_of("NO2") == pytest.approx(2.8)
    assert b.flow_of("N2O4") == pytest.approx(1.8)
    assert b.flow_of("NO2") == pytest.approx(1.2)


def test_splitter_ratio_sum_check():
    feed = Stream(["N2O4", "NO2"], flows={"N2O4": 10, "NO2": 0})
    a = Stream(["N2O4", "NO2"], name="a")
    b = Stream(["N2O4", "NO2"], name="b")
    with pytest.raises(ValueError):
        Splitter(feed, [a, b], ratios=[0.7, 0.7])


def test_recycle_with_splitter():
    comps = ["N2O4", "NO2"]
    feed = Stream(comps, flows={"N2O4": 10, "NO2": 0})
    rec = Stream(comps, name="rec")
    mixed = Stream(comps, name="mixed")
    rout = Stream(comps, name="rout")
    prod = Stream(comps, name="prod")
    rxn = Reaction({"N2O4": -1, "NO2": 2})
    p = Problem(
        [feed, rec, mixed, rout, prod],
        [
            Mixer([feed, rec], mixed),
            Reactor(mixed, rout, [rxn], key_component="N2O4", conversion=0.5),
            Splitter(rout, [prod, rec], ratios=[0.7, 0.3]),
        ],
    )
    sol = p.solve(bounds=(0, np.inf))
    assert sol.success
    assert rec.total_flow.eval() == pytest.approx(0.3 * rout.total_flow.eval())
    # 元素収支: 入 N = 出 N
    assert 2 * feed.flow_of("N2O4") == pytest.approx(
        2 * prod.flow_of("N2O4") + prod.flow_of("NO2")
    )


def test_generate_mermaid_shows_loop():
    comps = ["N2O4", "NO2"]
    feed = Stream(comps, flows={"N2O4": 10, "NO2": 0}, name="Feed")
    rec = Stream(comps, name="Recycle")
    mixed = Stream(comps, name="Mixed")
    rout = Stream(comps, name="ReactOut")
    prod = Stream(comps, name="Product")
    rxn = Reaction({"N2O4": -1, "NO2": 2})
    m = Mixer([feed, rec], mixed, name="M1")
    r = Reactor(mixed, rout, [rxn], key_component="N2O4", conversion=0.5, name="R1")
    sp = Splitter(rout, [prod, rec], ratios=[0.7, 0.3], name="SP1")
    src = generate_mermaid(Problem([feed, rec, mixed, rout, prod], [m, r, sp]))
    assert "flowchart" in src
    assert "feed_Feed" in src          # フィードノード
    assert "prod_Product" in src       # プロダクトノード
    assert "U_SP1 -->|Recycle| U_M1" in src  # 循環エッジ


def test_parse_pressure():
    assert parse_pressure(200000) == pytest.approx(200000)
    assert parse_pressure("2MPa") == pytest.approx(2e6)
    assert parse_pressure("2MPaG") == pytest.approx(2e6 + 101325)
    assert parse_pressure("1atm") == pytest.approx(101325)


def test_constrain_recovery():
    comps = ["CO", "H2O"]
    feed = Stream(comps, flows={"CO": 10, "H2O": 4})
    gas = Stream(comps, name="gas")
    liq = Stream(comps, name="liq")
    p = Problem([feed, gas, liq], [Separator(feed, [gas, liq])])
    # H2O は 75% を液へ、CO は 0%（全量ガス）
    p.constrain_recovery(feed, liq, {"H2O": 0.75, "CO": 0.0})
    sol = p.solve()
    assert sol.success
    assert liq.flow_of("H2O") == pytest.approx(3.0)
    assert liq.flow_of("CO") == pytest.approx(0.0)
    assert gas.flow_of("H2O") == pytest.approx(1.0)
    assert gas.flow_of("CO") == pytest.approx(10.0)


def test_gibbs_reactor_smr():
    pytest.importorskip("cantera")
    from chemflow2 import GibbsReactor

    species = ["CH4", "H2O", "CO", "CO2", "H2"]
    feed = Stream(species, flows={"CH4": 1.0, "H2O": 2.0, "CO": 0, "CO2": 0, "H2": 0},
                  condition=StreamCondition(T=850, P="0.1MPa"))
    out = Stream(species, name="out")
    g = GibbsReactor(feed, out, species=species, T=850, P="0.1MPa")
    sol = Problem([feed, out], [g]).solve()
    assert sol.success
    # 高温・S/C=2 では CH4 はほぼ完全に改質される
    assert out.flow_of("CH4") < 0.05
    # 元素 C 収支: 入 1 = 出（CO+CO2+CH4）
    assert out.flow_of("CO") + out.flow_of("CO2") + out.flow_of("CH4") == pytest.approx(1.0, abs=1e-3)


def test_to_excel(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    from chemflow2 import to_excel

    a = Stream(["N2", "H2"], name="1. A", order=1, flows={"N2": 1, "H2": 3},
               condition=StreamCondition(T=25, P="0.1MPaG", phase="gas"))
    b = Stream(["N2", "H2"], name="2. B", order=2, flows={"N2": 2, "H2": 1})
    out = Stream(["N2", "H2"], name="3. Out", order=3)
    p = Problem([a, b, out], [Mixer([a, b], out)])
    p.solve()

    path = tmp_path / "streams.xlsx"
    to_excel(p.streams, str(path), sheet="Streams")
    assert path.exists()

    ws = load_workbook(str(path))["Streams"]
    header = [c.value for c in ws[1]]
    assert header == ["Component", "MW", "1. A", "2. B", "3. Out"]
    # 総モル流量行の Out 列 = 3 + 4 = 7
    totals = {r[0].value: r for r in ws.iter_rows()}
    total_row = totals["total [mol/h]"]
    assert total_row[4].value == pytest.approx(7.0)


def test_dof_mismatch_raises():
    feed = Stream(["N2O4", "NO2"], flows={"N2O4": 10, "NO2": 0})
    out = Stream(["N2O4", "NO2"], name="out")
    # 未知 out に対しユニットも制約も無い → 自由度不足
    p = Problem([feed, out], [])
    with pytest.raises(SolveError):
        p.solve()
