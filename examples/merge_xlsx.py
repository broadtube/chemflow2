"""ユーティリティ: example_co2_removal.py が出す 3 つの xlsx を 1 つに結合する。

1 ケースにつき改質器・plant3・苛性ソーダ吸収塔の 3 フローシートが別々の xlsx に出る。
それを**列方向に連結**して、改質器から plant3 までを 1 枚のストリーム表で追えるようにする。

    co2removal_{label}_reformer.xlsx  ┐
    co2removal_{label}_plant3.xlsx    ├→ co2removal_{label}_merged.xlsx
    co2removal_{label}_caustic.xlsx   ┘

chemflow2 本体の機能ではなく example に置いてあるのは、この結合が
**example_co2_removal.py の出力レイアウトに固有**で汎用性が無いため。

────────────────────────────────────────────────────────────────────────
なぜ「最初から 1 ファイル」ではなく「後から結合」なのか
────────────────────────────────────────────────────────────────────────
最初から 1 ファイルにすると、3 つのフローシートを同一プロセスで同時に保持する必要が
あり、結合の都合が run_case の構造に侵入する。しかも**ケースをまたいだ結合ができない**
（28 runs は別プロセス呼び出しなので）。後段の独立スクリプトなら、あとから範囲を
変えて何度でも組み直せる。

────────────────────────────────────────────────────────────────────────
結合が「列の連結」だけで済む理由
────────────────────────────────────────────────────────────────────────
example_co2_removal.py は 3 ファイルとも **同じ basis（MERGE_BASIS）と同じ成分行
（MERGE_COMPONENTS）** で書き出している。したがって A 列（Component / T / P / Phase /
成分名 / total …）が 3 ファイルで完全に一致し、C 列以降をそのまま横に並べれば済む。

このスクリプトは xlsx のレイアウトを構文解析しない。**A 列が一致するかだけを検証**し、
ズレていれば黙って壊れる代わりに落ちる（`--force` で続行可）。行構造を揃えることと
結合が単純になることは、こうして噛み合っている。

────────────────────────────────────────────────────────────────────────
ストリーム番号
────────────────────────────────────────────────────────────────────────
個別ファイルの番号は**ローカルのまま**にしてある。各フローシートの Mermaid 図の丸数字と
一対一に対応しているので、振り直すと図と表が食い違うため（番号は Stream.name の文字列
"1. Steam1" にも埋まっている）。

グローバル番号は**結合シートだけが持つ属性**として、ここで BLOCKS のオフセットから
振る。結合シートのヘッダは 3 段:

    行1  ブロック帯      ── 改質器 ──        ── plant3 ──        ── T-101 吸収塔 ──
    行2  グローバル番号  101 102 …           301 302 …           401 402 …
    行3  ストリーム名    1. RG_feed 2. …     1. Steam1 2. …      1. GasIn 2. …

実行:
    python3 examples/merge_xlsx.py A_base_b_recycle_eta0.9
    python3 examples/merge_xlsx.py --all
    python3 examples/merge_xlsx.py --all --outdir /tmp/merged
要 openpyxl。chemflow2 本体には依存しない（xlsx を読んで xlsx を書くだけ）。
"""

import argparse
import glob
import os
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "output")

#: 結合の順序と、グローバル番号のオフセット。
#: プロセスの流れ（改質器 → plant3）に沿って並べ、吸収塔はその直後に置く。
#: 100 番刻みにしてあるのは、1 ブロックあたりのストリームが増えても衝突しないため。
#: 種別ごとに桁を分けておくと、結合シートだけ見ても出どころが分かる。
BLOCKS = [
    ("reformer", 100, "改質器 (pattern1)"),
    ("caustic",  200, "苛性ソーダ吸収塔"),
    ("plant3",   300, "MA プラント (plant3)"),
]

#: A 列に来るラベルのうち、値ではなく構造を表すもの（検証時の見出し）
_STRUCTURE_HEAD = "Component"


def source_paths(label: str, outdir: str) -> list[tuple[str, int, str, str]]:
    """(種別, オフセット, 表示名, パス) のリスト。存在するものだけ返す。

    baseline ケースには吸収塔が無いなど、ブロックの有無はケースによって変わる。
    """
    found = []
    for kind, offset, title in BLOCKS:
        path = os.path.join(outdir, f"co2removal_{label}_{kind}.xlsx")
        if os.path.exists(path):
            found.append((kind, offset, title, path))
    return found


def read_sheet(path: str) -> tuple[list, list[list]]:
    """(A列ラベルのリスト, データ列のリスト) を返す。

    データ列は「1 ストリーム = 1 リスト」で、先頭要素がストリーム名（ヘッダ行の値）。
    B 列（MW）は結合先で 1 本あればよいので、ここでは A 列側に含めて持ち回る。
    """
    ws = load_workbook(path, data_only=True).active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    labels = [(r[0], r[1]) for r in rows]          # (Component, MW) の 2 列
    n_cols = max(len(r) for r in rows)
    columns = []
    for c in range(2, n_cols):
        columns.append([r[c] if c < len(r) else None for r in rows])
    return labels, columns


def merge_label(label: str, outdir: str, *, force: bool = False) -> str | None:
    """1 ケース分を結合して書き出す。書いたパスを返す（対象が無ければ None）。"""
    sources = source_paths(label, outdir)
    if not sources:
        print(f"  {label}: 元ファイルが見つからない — スキップ")
        return None

    base_labels, blocks = None, []
    for kind, offset, title, path in sources:
        labels, columns = read_sheet(path)
        if base_labels is None:
            base_labels = labels
        elif labels != base_labels:
            # 行構造が違う = basis か成分行が揃っていない。連結すると値が縦にズレる。
            msg = (f"{label}: {kind} の行構造が他と一致しません。"
                   f" example_co2_removal.py の MERGE_BASIS / MERGE_COMPONENTS を"
                   f" 3 か所とも同じにして出力し直してください。")
            if not force:
                raise SystemExit(f"✗ {msg}")
            print(f"  ⚠ {msg}（--force のため続行）")
        blocks.append((kind, offset, title, columns))

    wb = Workbook()
    ws = wb.active
    ws.title = "merged"
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    # --- ヘッダ 3 段（ブロック帯 / グローバル番号 / ストリーム名）---
    band, numbers, names = ["", ""], ["", ""], list(base_labels[0])
    spans = []                                    # (開始列, 終了列, 表示名)
    for _kind, offset, title, columns in blocks:
        start = len(band) + 1
        for i, col in enumerate(columns, 1):
            band.append("")
            numbers.append(offset + i)
            names.append(col[0])                  # 元のストリーム名をそのまま使う
        spans.append((start, len(band), title))
    ws.append(band)
    ws.append(numbers)
    ws.append(names)
    for row in (1, 2, 3):
        for c in ws[row]:
            c.font = bold
            c.alignment = center
    for start, end, title in spans:
        ws.cell(row=1, column=start).value = title
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

    # --- 本体（A/B 列は先頭ブロックのものを使う。行構造は検証済み）---
    for r in range(1, len(base_labels)):
        row = list(base_labels[r])
        for _kind, _offset, _title, columns in blocks:
            row.extend(col[r] for col in columns)
        ws.append(row)
        if base_labels[r][0] and str(base_labels[r][0]).startswith(("total", "[")):
            ws.cell(row=ws.max_row, column=1).font = bold

    ws.freeze_panes = "C4"                        # ラベル列とヘッダを固定
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 8
    for i in range(3, len(names) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    path = os.path.join(outdir, f"co2removal_{label}_merged.xlsx")
    wb.save(path)
    n_streams = sum(len(c) for *_x, c in blocks)
    print(f"  {label}: {len(blocks)} ブロック / {n_streams} ストリーム → "
          f"{os.path.basename(path)}")
    return path


def discover_labels(outdir: str) -> list[str]:
    """output/ にある co2removal_*_{種別}.xlsx からケースラベルを拾う。"""
    kinds = "|".join(k for k, _o, _t in BLOCKS)
    pat = re.compile(rf"^co2removal_(.+)_({kinds})\.xlsx$")
    labels = {m.group(1) for f in glob.glob(os.path.join(outdir, "co2removal_*.xlsx"))
              if (m := pat.match(os.path.basename(f)))}
    return sorted(labels)


def main():
    ap = argparse.ArgumentParser(
        description="example_co2_removal.py の 3 つの xlsx を 1 つに結合する")
    ap.add_argument("labels", nargs="*",
                    help="ケースラベル（例 A_base_b_recycle_eta0.9）。省略時は --all が要る")
    ap.add_argument("--all", action="store_true", help="output/ にある全ケースを結合")
    ap.add_argument("--outdir", default=OUT, help=f"入出力ディレクトリ（既定 {OUT}）")
    ap.add_argument("--force", action="store_true",
                    help="行構造が一致しなくても結合を続行する（値が縦にズレる恐れ）")
    args = ap.parse_args()

    labels = discover_labels(args.outdir) if args.all else args.labels
    if not labels:
        ap.error("ラベルを指定するか --all を付けてください"
                 if not args.all else f"{args.outdir} に結合対象がありません")

    print(f"結合対象 {len(labels)} ケース（{args.outdir}）")
    written = [p for lb in labels if (p := merge_label(lb, args.outdir, force=args.force))]
    print(f"\n{len(written)} / {len(labels)} ケースを書き出しました")
    return 0 if written else 1


if __name__ == "__main__":
    sys.exit(main())
